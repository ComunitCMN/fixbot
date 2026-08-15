"""
Выгрузка базы застройщика в Excel.

Решение целиком — в РЕШЕНИЯ.md, раздел «Выгрузка базы застройщика
в Excel». Здесь закреплено главное его свойство и структура файла.

**Выгрузка только читает.** Она не пишет ни в amoCRM, ни в базу и не
трогает схему. Это не деталь реализации, а то, ради чего задачу вообще
можно делать на живом проекте: при таком ограничении сломать ею
работающие фиксации технически нечем. Три первых теста стерегут ровно
это, и «тест устарел» про них — неправильный ответ.

Зеркало здесь ни при чём совсем: выгрузка обходит amoCRM сама. Зеркало
живёт своей жизнью и ловит дубли, и эта задача его не касается.
"""

import inspect

import pytest

from amo import AmoClient
from db import Db

import export_base as ex


# ===================== заглушка amoCRM =====================

class FakeAuth:
    subdomain = "test"

    async def token(self):
        return "t"

    async def close(self):
        pass


PIPELINES = [
    {"id": 1, "name": "Крым Лиды", "sort": 10, "statuses": [
        {"id": 11, "name": "Неразобранное", "sort": 10, "type": 1},
        {"id": 12, "name": "Взят в работу", "sort": 20, "type": 0},
        {"id": 13, "name": "Показ проведён", "sort": 30, "type": 0},
        {"id": 142, "name": "Успешно реализовано", "sort": 10000, "type": 0},
        {"id": 143, "name": "Закрыто и не реализовано", "sort": 11000,
         "type": 0},
    ]},
    {"id": 2, "name": "Агентские Клиенты", "sort": 20, "statuses": [
        {"id": 21, "name": "Первичный контакт", "sort": 10, "type": 0},
        {"id": 142, "name": "Успешно реализовано", "sort": 10000, "type": 0},
        {"id": 143, "name": "Закрыто и не реализовано", "sort": 11000,
         "type": 0},
    ]},
]

LEADS = [
    {"id": 100, "pipeline_id": 1, "status_id": 12, "name": "Сделка A",
     "created_at": 1700000000, "updated_at": 1700000000,
     "_embedded": {"contacts": [{"id": 900}]}},
    {"id": 101, "pipeline_id": 1, "status_id": 13, "name": "Сделка B",
     "created_at": 1700000000, "updated_at": 1700000000,
     "_embedded": {"contacts": [{"id": 901}]}},
    {"id": 102, "pipeline_id": 1, "status_id": 142, "name": "Сделка C",
     "created_at": 1700000000, "updated_at": 1700000000,
     "_embedded": {"contacts": [{"id": 902}]}},
    {"id": 103, "pipeline_id": 1, "status_id": 143, "name": "Сделка D",
     "created_at": 1700000000, "updated_at": 1700000000,
     "_embedded": {"contacts": [{"id": 903}],
                   "loss_reason": [{"id": 7, "name": "Дорого"}]}},
    {"id": 104, "pipeline_id": 1, "status_id": 11, "name": "Сделка E",
     "created_at": 1700000000, "updated_at": 1700000000,
     "_embedded": {"contacts": [{"id": 904}]}},
    {"id": 200, "pipeline_id": 2, "status_id": 21, "name": "Сделка F",
     "created_at": 1700000000, "updated_at": 1700000000,
     "_embedded": {"contacts": [{"id": 900}]}},
]

CONTACTS = [
    {"id": 900, "name": "Алмаз", "custom_fields_values": [
        {"field_code": "PHONE", "values": [{"value": "+7 917 932-10-03"}]}]},
    {"id": 901, "name": "Фаиль", "custom_fields_values": [
        {"field_code": "PHONE", "values": [{"value": "+79274248100"}]}]},
    {"id": 902, "name": "Дамир", "custom_fields_values": []},
    {"id": 903, "name": "Азат", "custom_fields_values": [
        {"field_code": "PHONE", "values": [{"value": "+79950080757"}]}]},
    {"id": 904, "name": "Айзат", "custom_fields_values": [
        {"field_code": "PHONE", "values": [{"value": "+79534081924"}]}]},
]


class FakeAmo(AmoClient):
    """
    Настоящий AmoClient с подменённым транспортом: каждый запрос
    записывается вместе с методом. Именно по этой записи проверяется,
    что выгрузка ничего не меняет.
    """

    def __init__(self, pipelines=None, leads=None, contacts=None):
        super().__init__(FakeAuth())
        self.calls: list[dict] = []
        self._pipelines = PIPELINES if pipelines is None else pipelines
        self._leads = LEADS if leads is None else leads
        self._contacts = CONTACTS if contacts is None else contacts

    async def _request(self, method, url, **kw):
        params = kw.get("params") or {}
        self.calls.append({"method": method, "url": url, "params": params,
                           "json": kw.get("json")})

        if url == "/api/v4/leads/pipelines":
            return {"_embedded": {"pipelines": [
                {"id": p["id"], "name": p["name"], "sort": p["sort"],
                 "_embedded": {"statuses": p["statuses"]}}
                for p in self._pipelines]}}

        # Постранично: вторая страница пустая — обход на ней остановится.
        page = params.get("page", 1) if isinstance(params, dict) else 1
        if url == "/api/v4/leads":
            items = self._leads if page == 1 else []
            return {"_embedded": {"leads": items}}
        if url == "/api/v4/contacts":
            items = self._contacts if page == 1 else []
            return {"_embedded": {"contacts": items}}
        return {}


async def _collected(amo=None):
    amo = amo or FakeAmo()
    return amo, await ex.collect(amo)


# ===================== главное: выгрузка только читает =====================

@pytest.mark.asyncio
async def test_export_only_reads_from_amocrm():
    """
    Ни одного изменяющего запроса. Появится POST, PATCH или DELETE —
    значит выгрузка начала что-то менять в CRM застройщика, и это
    отмена решения, а не устаревший тест.
    """
    amo, _ = await _collected()

    assert amo.calls, "выгрузка не сходила в amoCRM вовсе"
    methods = {c["method"] for c in amo.calls}
    assert methods == {"GET"}, f"выгрузка меняет данные в amoCRM: {methods}"
    assert all(c["json"] is None for c in amo.calls), \
        "выгрузка отправила тело запроса — значит что-то записывает"


@pytest.mark.asyncio
async def test_export_does_not_touch_the_database(tmp_path):
    """
    База после выгрузки побайтно та же: ни строк, ни схемы, ни миграций.
    Снимок делается полным дампом — он ловит и данные, и структуру.
    """
    db = Db(tmp_path / "e.db")
    db.replace_pipelines([{"id": 1, "name": "Крым Лиды", "statuses": [
        {"id": 12, "name": "Взят в работу", "sort": 20, "type": 0}]}])
    db.replace_contacts([{"id": 900, "name": "Алмаз",
                          "phones": ["+79179321003"], "created_at": 1}])

    before = list(db.conn.iterdump())

    amo, data = await _collected()
    ex.build_workbook(**data)

    assert list(db.conn.iterdump()) == before, \
        "выгрузка изменила базу — а она должна только читать"


def test_export_module_never_writes_anywhere():
    """
    Страховка на будущее: в модуле выгрузки нет ни записи в базу,
    ни изменяющих запросов. Ловит правку, которую первые два теста
    не увидят, — например, кэш «чтобы не ходить в CRM дважды».
    """
    src = inspect.getsource(ex)

    for forbidden in ("INSERT", "UPDATE ", "DELETE", "CREATE TABLE",
                      "ALTER TABLE", "commit()"):
        assert forbidden not in src, f"выгрузка пишет в базу: {forbidden}"

    for method in ('"POST"', "'POST'", '"PATCH"', "'PATCH'",
                   '"DELETE"', "'DELETE'"):
        assert method not in src, f"выгрузка меняет amoCRM: {method}"

    assert "import db" not in src and "from db " not in src, \
        "выгрузка дотянулась до базы — она не должна её знать"


# ===================== структура файла =====================

@pytest.mark.asyncio
async def test_sheet_per_pipeline():
    """Лист на воронку — как договорились."""
    _, data = await _collected()
    names = ex.sheet_names(data["pipelines"])

    assert names == ["Крым Лиды", "Агентские Клиенты"]


@pytest.mark.asyncio
async def test_won_on_top_and_lost_at_the_bottom():
    """
    Сверху успешно реализованные, внизу отказы. Смотреть такой файл
    начинают сверху, и первым должно идти то, ради чего всё делалось.
    """
    _, data = await _collected()
    stages = [s["name"] for s in ex.order_statuses(data["pipelines"][0]["statuses"])]

    assert stages[0] == "Успешно реализовано"
    assert stages[-1] == "Закрыто и не реализовано"


@pytest.mark.asyncio
async def test_working_stages_go_in_reverse_pipeline_order():
    """Ближе к сделке — выше."""
    _, data = await _collected()
    stages = [s["name"] for s in ex.order_statuses(data["pipelines"][0]["statuses"])]

    assert stages == ["Успешно реализовано", "Показ проведён", "Взят в работу",
                      "Неразобранное", "Закрыто и не реализовано"]


@pytest.mark.asyncio
async def test_rows_follow_the_stage_order():
    _, data = await _collected()
    rows = ex.sheet_rows(data["pipelines"][0], data["leads"], data["contacts"])
    stages = [r[ex.COLUMNS.index("Этап")] for r in rows]

    assert stages == ["Успешно реализовано", "Показ проведён", "Взят в работу",
                      "Неразобранное", "Закрыто и не реализовано"]


# ===================== содержимое строк =====================

@pytest.mark.asyncio
async def test_loss_reason_lands_in_its_column():
    """
    Причина отказа приходит из amoCRM тем же запросом, что и этапы.
    В зеркале её нет и не будет — проверено, в РЕШЕНИЯ.md записано.
    """
    _, data = await _collected()
    rows = ex.sheet_rows(data["pipelines"][0], data["leads"], data["contacts"])
    by_stage = {r[ex.COLUMNS.index("Этап")]: r for r in rows}
    reason = ex.COLUMNS.index("Причина отказа")

    assert by_stage["Закрыто и не реализовано"][reason] == "Дорого"
    # У живых сделок причины отказа быть не может — иначе колонка врёт.
    assert by_stage["Взят в работу"][reason] == ""
    assert by_stage["Успешно реализовано"][reason] == ""


@pytest.mark.asyncio
async def test_client_name_and_phone_come_from_the_contact():
    _, data = await _collected()
    rows = ex.sheet_rows(data["pipelines"][0], data["leads"], data["contacts"])
    by_stage = {r[ex.COLUMNS.index("Этап")]: r for r in rows}
    row = by_stage["Взят в работу"]

    assert row[ex.COLUMNS.index("Клиент")] == "Алмаз"
    assert row[ex.COLUMNS.index("Телефон")] == "+7 917 932-10-03"


@pytest.mark.asyncio
async def test_lead_without_a_phone_still_appears():
    """
    Контакт без телефона — обычное дело, и терять из-за этого сделку
    нельзя: в файле она нужна, просто без номера.
    """
    _, data = await _collected()
    rows = ex.sheet_rows(data["pipelines"][0], data["leads"], data["contacts"])
    by_stage = {r[ex.COLUMNS.index("Этап")]: r for r in rows}

    assert by_stage["Успешно реализовано"][ex.COLUMNS.index("Клиент")] == "Дамир"
    assert by_stage["Успешно реализовано"][ex.COLUMNS.index("Телефон")] == ""


@pytest.mark.asyncio
async def test_lead_without_a_contact_is_not_lost():
    amo = FakeAmo(leads=[{"id": 300, "pipeline_id": 1, "status_id": 12,
                          "name": "Ничей", "created_at": 1700000000,
                          "updated_at": 1700000000, "_embedded": {}}])
    _, data = await _collected(amo)
    rows = ex.sheet_rows(data["pipelines"][0], data["leads"], data["contacts"])

    assert len(rows) == 1
    assert rows[0][ex.COLUMNS.index("Клиент")] == ""


@pytest.mark.asyncio
async def test_lead_in_an_unknown_stage_is_not_dropped():
    """
    Этап могли завести после последнего чтения справочника. Молча терять
    такую сделку нельзя — файл станет неполным, и никто не заметит.
    """
    amo = FakeAmo(leads=[{"id": 301, "pipeline_id": 1, "status_id": 999999,
                          "name": "Новый этап", "created_at": 1700000000,
                          "updated_at": 1700000000, "_embedded": {}}])
    _, data = await _collected(amo)
    rows = ex.sheet_rows(data["pipelines"][0], data["leads"], data["contacts"])

    assert len(rows) == 1


# ===================== сам файл =====================

@pytest.mark.asyncio
async def test_workbook_opens_and_has_the_sheets():
    from io import BytesIO

    from openpyxl import load_workbook

    _, data = await _collected()
    wb = load_workbook(BytesIO(ex.build_workbook(**data)))

    assert wb.sheetnames == ["Крым Лиды", "Агентские Клиенты"]
    first = wb["Крым Лиды"]
    assert [c.value for c in first[1]] == list(ex.COLUMNS)
    assert first.max_row == 1 + 5


def test_long_pipeline_name_fits_the_sheet_title():
    """
    Excel не берёт длинные названия и часть знаков. Файл, который
    не открывается, — худшая поломка: она видна только у оператора.
    """
    names = ex.sheet_names([
        {"id": 1, "name": "Очень длинное название воронки [Крым] / 2026",
         "statuses": []},
        {"id": 2, "name": None, "statuses": []},
    ])

    assert len(names) == 2, "воронка потерялась вместе с названием"
    assert all(len(n) <= 31 for n in names)
    assert not any(set(n) & set("[]:*?/\\") for n in names)
    assert all(n for n in names), "лист без названия Excel не примет"


def test_sheets_with_the_same_name_do_not_collide():
    names = ex.sheet_names([{"id": 1, "name": "Лиды", "statuses": []},
                            {"id": 2, "name": "Лиды", "statuses": []}])

    assert len(set(names)) == 2


def test_huge_pipeline_continues_on_the_next_sheet():
    """
    У крупного застройщика сделок в сотни раз больше. Лист Excel
    вмещает миллион с небольшим строк; на большем openpyxl падает,
    и файл не приходит вовсе.
    """
    assert ex.SHEET_ROW_LIMIT <= 1048576
    chunks = ex.split_rows([[i] for i in range(ex.SHEET_ROW_LIMIT + 5)])

    assert len(chunks) == 2
    assert sum(len(c) for c in chunks) == ex.SHEET_ROW_LIMIT + 5


# ===================== кнопка =====================

def test_only_the_operator_sees_the_export_button():
    """
    Файл с полной базой застройщика — самый ценный артефакт во всей
    системе. Владельцу кнопку не показываем, как и воронки.
    """
    import menu as mn

    tech = [b.text for r in mn.tech_menu().inline_keyboard for b in r]
    assert any("Выгрузк" in x for x in tech)

    for flag in (False, True):
        owner = [b.text for r in mn.main_menu(mn.OWNER, flag).inline_keyboard
                 for b in r]
        assert not any("Выгрузк" in x for x in owner)
        assert "🔧 Техническое" not in owner


class _Msg:
    async def edit_text(self, text, **kw):
        self.text = text

    async def answer(self, text, **kw):
        self.text = text


class _Press:
    """Нажатие кнопки: и настоящее, и подделанное в обход меню."""

    def __init__(self, user_id: int, data: str = "m:export"):
        self.from_user = type("U", (), {"id": user_id})()
        self.data = data
        self.message = _Msg()
        self.alerts: list[str] = []

    async def answer(self, text="", show_alert=False):
        self.alerts.append(text)


@pytest.fixture
def pressing(monkeypatch):
    """
    Готовит бота к нажатию кнопки: оператор — 1, владелец — 2, его
    сотрудник — 3. Настоящую выгрузку не запускаем, только смотрим,
    дошло ли до неё дело.
    """
    import bot as b

    started: list[str] = []

    def spy(coro):
        started.append(getattr(coro, "__qualname__", str(coro)))
        coro.close()
        return None

    monkeypatch.setattr(b.asyncio, "create_task", spy)
    monkeypatch.setattr(b.cfg, "operator_ids", {1})
    monkeypatch.setattr(b.cfg, "owner_ids", {2})
    monkeypatch.setattr(b.db, "is_staff", lambda uid: uid == 3)

    async def press(user_id: int):
        started.clear()
        cb = _Press(user_id)
        await b.cb_menu(cb)
        return cb, list(started)

    return press


@pytest.mark.asyncio
async def test_owner_cannot_start_the_export_even_past_the_menu(pressing):
    """
    Кнопки у владельца нет, но кнопка — это не защита: нажатие можно
    подделать. Отказать должен обработчик. Сотрудник владельца —
    маркетолог, РОП — идёт по той же роли, и ему тоже нельзя.
    """
    for user_id, who in ((2, "владелец"), (3, "сотрудник владельца"),
                         (99, "посторонний")):
        cb, started = await pressing(user_id)

        assert not started, f"{who} запустил выгрузку базы застройщика"
        assert cb.alerts == ["Недоступно"], f"{who} получил не отказ"


@pytest.mark.asyncio
async def test_operator_can_start_the_export(pressing):
    """Обратная сторона: у оператора кнопка обязана работать."""
    cb, started = await pressing(1)

    assert started, "у оператора выгрузка не запустилась"
    assert "Выгрузка базы" in cb.message.text


def test_export_runs_in_the_background():
    """
    Полный обход amoCRM — минуты. Ответить надо сразу, иначе Telegram
    оборвёт кнопку по таймауту и это прочтётся как поломка.
    """
    import bot as b

    src = inspect.getsource(b.cb_menu)
    start = src.index('section == "export"')
    assert "create_task" in src[start:start + 900], \
        "выгрузка выполняется прямо в обработчике — кнопка отвалится"


def test_export_is_only_by_hand():
    """
    Ни расписания, ни автоматических выгрузок: файл с полной базой
    застройщика не должен появляться сам.
    """
    import bot as b

    assert "export_loop" not in inspect.getsource(b)
    # Фоновые задачи бот заводит при старте. Выгрузки среди них быть
    # не должно: она запускается только нажатием оператора.
    assert "export" not in inspect.getsource(b.main)
